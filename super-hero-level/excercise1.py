# Workbooks
# Creating a new workbook
from openpyxl import Workbook

workbook = Workbook()  # Creating a new workbook object, by initializing the Workbook() class

file_path = "D:\\ExcelApp\\MyCompanyStaff.xlsx"  # Setting the path to (location of) the new Excel workbook

workbook.save(file_path)  # Saving the workbook

# Loading the Excel workbook .xlsx file
workbook = openpyxl.load_workbook("/Users/gabriel/Downloads/Employees.xlsx")

# Getting the basic properties of the workbook
workbook.properties

# Getting the sheets in the workbook
workbook.worksheets

# Getting the sheet names in a workbook, as a list of strings
workbook.sheetnames

# Getting the active sheet in the workbook
workbook.active

# Referencing a sheet by its name
workbook['EmployeeData']

sheet = workbook['EmployeeData']

# Creating a new sheet in the workbook
workbook.create_sheet('TestSheet')
workbook.save("/Users/gabriel/Downloads/Employees.xlsx")  # Saving the workbook

# Removing a sheet from the workbook
sheet = workbook['TestSheet']
workbook.remove(sheet)

# or

del workbook['TestSheet']

workbook.save("/Users/gabriel/Downloads/Employees.xlsx")  # Saving the workbook

# Closing a workbook
workbook.close()
